"""
    随机生成产品特征数据，需要和市场订单数据的结果关联匹配
    Created By: 段小飞
    Creation Date: 2024-4-24
"""
import random
import pandas as pd
import datetime
from datetime import date
from faker import Faker
from collections import OrderedDict
from time import sleep
from datetime import datetime, timedelta
from tqdm import tqdm
from openpyxl import load_workbook
import re

def main():
    # 初始化设置
    fake = Faker('zh-CN')  # 指定用中文
    num = 100000  # 初始化模拟数据生成数量
    filename = 'ccjz_ddsj.xlsx'  # 定义文件名
    sheetName = 'ccjz_ddsj'  # 定义工作表名称

    # 申明字段
    customer_id = []            # 1、购买者ID
    shipping_country = []       # 2、购买者国家
    comment_date = []           # 3、评论时间
    size = []                   # 4、大小
    color = []                  # 5、颜色
    material = []               # 6、材质
    comment = []                # 7、评论内容
    star_level = []             # 8、星级

    print("----------1、读取用户信息表----------")
    wb = load_workbook(filename)
    sheet = wb.get_sheet_by_name(sheetName)  # 通过工作表名获取一个工作表对象
    rows = [val for val in sheet.rows]  # 先用列表推导式生产包含每一列中所有单元格的元组的列表
    maxRowNum = sheet.max_row  # 获取工作表中的最大行号

    print("----------开始生成数据，祈祷不报错----------")
    ii = 0
    for i in tqdm(range(1, maxRowNum)):
        row = rows[i]

        """
            这里是写每条数据生成的逻辑
        """
        # 1、购买者ID：customer_id
        customer_id_value = ''
        customer_id_value = row[18].value  # 获取用户ID
        customer_id.append(customer_id_value)
        # 2、购买者国家：shipping_country
        shipping_country_value = ''
        shipping_country_value = row[19].value  # 购买者国家
        shipping_country.append(shipping_country_value)
        # 3、评论时间：comment_date
        comment_date_value = ''
        order_date_value = row[4].value   # 订单日期
        random_days = random.randint(0, 30)
        comment_date_value = order_date_value + timedelta(days=random_days)
        comment_date.append(comment_date_value)
        # 4、大小：size
        size_value = ''
        material_value = ''
        product_type_value = row[3].value # 商品类型
        product_name_value = row[2].value  # 商品名称
        if product_type_value == '宠物咀嚼玩具':
            size_value = re.search(r'(\d+cm)', product_name_value).group(1)
            material_value = '谷物水果淀粉混合材质'
        elif product_type_value == '宠物如厕训练垫':
            size_value = re.search(r'(\d+片)', product_name_value).group(1)
            material_value = '无纺布及聚酯纤维材质'
        elif product_type_value == '宠物洗发露':
            size_value = re.search(r'(\d+ml)', product_name_value).group(1)
            material_value = '植物提取物及温和清洗剂'
        elif product_type_value == '伸缩式狗绳':
            size_value = re.search(r'(\d+m)', product_name_value).group(1)
            material_value = '尼龙结合金属'
        size.append(size_value)
        material.append(material_value)
        # 5、颜色：color
        color_value = ''
        color_dict = OrderedDict([('咖啡色', 0.17), ('灰色', 0.1), ('黄色', 0.1), ('红色', 0.1), ('蓝色', 0.1), ('粉色', 0.1), ('迷彩色', 0.13), ('橘色', 0.1), ('绿色', 0.1)])
        color_value = fake.random_element(color_dict)
        color.append(color_value)
        # 6、材质：material

        # 7、评论内容：comment
        comment_value = ''
        # 定义一些常用的评论片段
        comment_fragments = [
            "This product is ",
            "I really like the ",
            "The quality of the ",
            "The customer service was ",
            "I would definitely recommend ",
            "The price is very ",
            "The packaging is ",
            "It works great for ",
            "I use it for ",
            "Overall, I am ",
        ]

        # 定义一些形容词和副词，用于增加评论的多样性
        adjectives = [
            "amazing",
            "fantastic",
            "incredible",
            "wonderful",
            "excellent",
            "terrible",
            "poor",
            "great",
            "good",
            "bad",
        ]
        adverbs = [
            "really",
            "extremely",
            "very",
            "quite",
            "fairly",
            "somewhat",
        ]
        nouns = [
            "product",
            "service",
            "quality",
            "price",
            "packaging",
            "experience",
        ]
        # 随机选择一个评论片段、形容词、副词和名词来生成评论
        def generate_random_comment():
            fragment = random.choice(comment_fragments)
            adj = random.choice(adjectives)
            adv = random.choice(adverbs)
            noun = random.choice(nouns)
            # 根据所选的片段构造评论
            if "The " in fragment:
                comment = f"{fragment}{adj}."
            elif fragment.endswith(" is "):
                comment = f"{fragment}{adj}."
            elif fragment.endswith(" for "):
                comment = f"{fragment}{noun} and it's {adv} {adj}."
            else:
                comment = f"{fragment} {adv} {adj}."
            return comment
        comment_value = generate_random_comment()
        comment.append(comment_value)
        # 8、星级：star_level
        star_level_value = ''
        if 'amazing' in comment_value:
            star_level_value = '5星'
        if 'fantastic' in comment_value:
            star_level_value = '5星'
        if 'wonderful' in comment_value:
            star_level_value = '4星'
        if 'incredible' in comment_value:
            star_level_value = '4星'
        if 'excellent' in comment_value:
            star_level_value = '3星'
        if 'great' in comment_value:
            star_level_value = '3星'
        if 'good' in comment_value:
            star_level_value = '2星'
        if 'terrible' in comment_value:
            star_level_value = '1星'
        if 'poor' in comment_value:
            star_level_value = '1星'
        if 'bad' in comment_value:
            star_level_value = '1星'

        star_level.append(star_level_value)


        # 把前面根据逻辑生成的数据，赋值到data这个临时的数据集里面
        data = {
            '购买者ID': customer_id,
            '购买者国家': shipping_country,
            '评论时间': comment_date,
            '大小': size,
            '颜色': color,
            '材质': material,
            '评论内容': comment,
            '星级': star_level
        }

    # 把生成的data的数据集，导出成Excel文件
    pd.set_option('display.width', None)  # 设置字符显示无限制
    pd.set_option('display.max_rows', None)  # 设置行数显示无限制
    users = pd.DataFrame.from_dict(data)
    # print(users)
    users.to_csv("ccjz_ddsj_product.xlsx", index=True) # 指定生成文件的名字和格式
    print("\n" + "----------执行结束，万幸----------")

if __name__ == '__main__':
    main()