"""
        自动生成用户事件数据
        案例：BCW
        Created By：xujian
        Created Time：2022-11-29
        本案例只有两次操作，分别是：
        1、通过数字营销系统自动给不同类型的用户，发送不同的促销信息，对发送成功的消息，以及发送失败的消息，系统都有事件记录
        2、用户是否点击发送的促销消息，系统有事件记录
        3、对于发送的促销消息之后，用户是否有产生购买行为，系统有事件记录
"""
import datetime
import random
import pandas as pd
from datetime import date
from faker import Faker
from tqdm import tqdm
from openpyxl import load_workbook
from collections import OrderedDict

def main():
    fake = Faker('zh-CN')  # 指定用中文
    num = 50000  # 初始化模拟数据生成数量
    filename = 'bcw_user.xlsx'  # 定义文件名
    sheetName = 'bcw_user'  # 定义工作表名称
    """  初始化事件属性字段以及字段对应的数据字典  """
    action_id = []      # 1、用户事件ID
    user_id = []        # 2、客户ID
    action_type = []    # 3、事件类型：发送促销信息，购买产品
    action_detail = []  # 4、事件动作：对于发送促销消息，动作有：发送成功，发送失败。对于购买产品，动作有：香辣味5袋、火锅味5袋、香辣味2袋+火锅味2袋+椒麻味/卤香味1袋随机、口味盲盒5袋、椒麻味2袋+卤香味2袋+香辣味/火锅味1袋随机
    action_time = []    # 5、事件时间
    print("----------1、读取用户信息表----------")
    wb = load_workbook(filename)
    sheet = wb.get_sheet_by_name(sheetName)  # 通过工作表名获取一个工作表对象
    rows = [val for val in sheet.rows]  # 先用列表推导式生产包含每一列中所有单元格的元组的列表
    maxRowNum = sheet.max_row  # 获取工作表中的最大行号
    print("----------2、开始生成数据，祈祷不报错----------")
    for i in tqdm(range(0, num)):
        """  1、生成action_id  """
        action_id_value = fake.random_int(min=10000000, max=99999999, step=1)
        action_id.append(action_id_value)
        """  2、生成user_id，读取之前生成的用户数据，随机取数，形成用户ID  """
        indexs = random.sample(range(0, maxRowNum - 1), 1) # 获取0到最大行号中的1个整数
        for index in indexs:
            row = rows[index]
            user_id_value = row[0].value
        user_id.append(user_id_value)
        """  3、生成action_time  """
        action_time_value = fake.date_time_between(start_date=date(2022, 9, 10), end_date=date(2022, 9, 30))
        action_time.append(action_time_value)
        """  4、生成action_type  """
        if action_time_value < datetime.datetime(2022, 9, 10):
            action_type_dict = OrderedDict([('发送促销信息', 0.85), ('购买产品', 0.15)])
            action_type_value = fake.random_element(action_type_dict)
            action_type.append(action_type_value)
        elif action_time_value > datetime.datetime(2022, 9, 30):
            action_type_dict = OrderedDict([('发送促销信息', 0.80), ('购买产品', 0.20)])
            action_type_value = fake.random_element(action_type_dict)
            action_type.append(action_type_value)
        else:
            action_type_dict = OrderedDict([('发送促销信息', 0.55), ('购买产品', 0.45)])
            action_type_value = fake.random_element(action_type_dict)
            action_type.append(action_type_value)
        """  5、生成action_detail  """
        if action_type_value == "发送促销信息":
            action_detail_dict = OrderedDict([('发送成功', 0.75), ('发送失败', 0.25)])
            action_detail_value = fake.random_element(action_detail_dict)
            action_detail.append(action_detail_value)
        elif action_type_value == "购买产品":
            action_detail_dict = OrderedDict([('香辣味5袋', 0.35), ('火锅味5袋', 0.15), ('口味盲盒5袋', 0.15), ('香辣味2袋+火锅味2袋+椒麻味/卤香味1袋随机', 0.20), ('椒麻味2袋+卤香味2袋+香辣味/火锅味1袋随机', 0.15)])
            action_detail_value = fake.random_element(action_detail_dict)
            action_detail.append(action_detail_value)

        data = {
            '用户事件ID': action_id,
            '客户ID': user_id,
            '事件类型': action_type,
            '事件动作': action_detail,
            '事件时间': action_time
        }
    pd.set_option('display.width', None)  # 设置字符显示无限制
    pd.set_option('display.max_rows', None)  # 设置行数显示无限制
    user_actions = pd.DataFrame.from_dict(data)
    # print(user_actions)
    user_actions.to_csv("bcw_user_actions.xlsx", index=True)
    print("\n" + "----------执行结束，万幸----------")

if __name__ == '__main__':
    main()