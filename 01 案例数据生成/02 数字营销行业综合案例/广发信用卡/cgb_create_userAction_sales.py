"""
        自动生成营销行动之后的用户事件数据
        Created By：段小飞
        Created Time：2023-09-13
        update Time：2024-04-11
"""

import random
import datetime
import faker
import pandas as pd
from faker import Faker
from tqdm import tqdm
from openpyxl import load_workbook
from collections import OrderedDict

def main():
    fake = Faker('zh-CN')  # 指定用中文
    num = 1000  # 初始化模拟数据生成数量
    filename = 'cgb_user.xlsx'  # 定义文件名
    sheetName = 'cgb_user'  # 定义工作表名称

    """  初始化事件属性字段以及字段对应的数据字典  """
    action_id = []      # 1、事件ID
    action_time = []    # 2、事件发生时间
    user_id = []        # 2、客户ID
    action_type = []    # 3、事件类型：餐饮、大额消费、电商、购物、旅游、其他
    amount = []         # 4、消费金额
    shop_no = []        # 5、收款方ID

    print("----------1、读取用户信息表----------")
    wb = load_workbook(filename)
    sheet = wb.get_sheet_by_name(sheetName)  # 通过工作表名获取一个工作表对象
    rows = [val for val in sheet.rows]  # 先用列表推导式生产包含每一列中所有单元格的元组的列表
    maxRowNum = sheet.max_row  # 获取工作表中的最大行号
    print("----------2、开始生成数据，祈祷不报错----------")
    ii = 0
    for i in tqdm(range(0, num)):
        ii = ii + 1
        """  1、生成user_id，读取之前生成的用户数据，随机取数，形成用户ID  """
        indexs = random.sample(range(1, maxRowNum - 1), 1)  # 获取0到最大行号中的1个整数
        for index in indexs:
            row = rows[index]
            user_id_value = row[0].value  # 获取"用户ID"信息
            have_creditcard = row[10].value  # 获取"是否信用卡用户"信息

        """  如果是无卡客户，则销售流程节点为四个，需要准备对应的节点事件数据  """
        if have_creditcard == '否':
            """  第1步：生成 发送短信 的事件数据（所有符合条件的用户都发送短信） """
            # 1、生成action_id
            action_id.append(ii)
            # 2、生成action_time
            action_time_value = fake.date_time_between(start_date=datetime.datetime(2023, 9, 1, 9, 30, 00), end_date=datetime.datetime(2023, 9, 1, 10, 00, 00))
            action_time.append(action_time_value)
            # 3、赋值user_id
            user_id.append(user_id_value)
            # 4、生成action_type
            action_type.append('无卡用户+发送短信')
            # 5、生成amount
            amount.append(0)
            # 6、生成shop_no
            shop_no.append('无')

            """  第2步：生成 收到短信打开链接 的事件数据 """
            # 1、生成action_id
            # 2、生成action_time
            action_time_value = fake.date_time_between(start_date=datetime.datetime(2023, 9, 1, 9, 30, 00), end_date=datetime.datetime(2023, 9, 1, 10, 30, 00))
            # 3、赋值user_id
            # 4、生成action_type
            action_type_dict = OrderedDict([('无卡用户+收到短信打开链接', 0.50), ('无', 0.50)])
            action_type_value = fake.random_element(action_type_dict)
            # 5、生成amount
            # 6、生成shop_no
            if action_type_value != '无':
                ii = ii + 1
                action_id.append(ii)
                action_time.append(action_time_value)
                user_id.append(user_id_value)
                action_type.append(action_type_value)
                amount.append(0)
                shop_no.append('无')
            if action_type_value != '无':
                """  第3步：生成 关注公众号 的事件数据 """
                # 1、生成action_id
                # 2、生成action_time
                # action_time_value = fake.date_time_between(start_date=datetime.datetime(2023, 9, 1, 10, 30, 00), end_date=datetime.datetime(2023, 9, 3, 23, 30, 00))
                action_time_value = fake.date_between(start_date=datetime.date(2023, 9, 1), end_date=datetime.date(2023, 9, 3))
                hour_temp = fake.random_int(min=10, max=22, step=1)
                if hour_temp < 10:
                    hour_temp2 = '0' + str(hour_temp)
                else:
                    hour_temp2 = str(hour_temp)
                min_temp = fake.random_int(min=0, max=59, step=1)
                if min_temp < 10:
                    min_temp2 = '0' + str(min_temp)
                else:
                    min_temp2 = str(min_temp)
                sec_temp = fake.random_int(min=0, max=59, step=1)
                if sec_temp < 10:
                    sec_temp2 = '0' + str(sec_temp)
                else:
                    sec_temp2 = str(sec_temp)
                action_time_value = str(action_time_value) + ' ' + hour_temp2 + ':' + min_temp2 + ':' + sec_temp2
                # 3、赋值user_id
                # 4、生成action_type
                action_type_dict = OrderedDict([('无卡用户+关注公众号', 0.75), ('无', 0.25)])
                action_type_value = fake.random_element(action_type_dict)
                # 5、生成amount
                # 6、生成shop_no
                if action_type_value != '无':
                    ii = ii + 1
                    action_id.append(ii)
                    action_time.append(action_time_value)
                    user_id.append(user_id_value)
                    action_type.append(action_type_value)
                    amount.append(0)
                    shop_no.append('无')
                if action_type_value != '无':
                    """  第4步：生成 收到资料打开链接 的事件数据 """
                    # 1、生成action_id
                    # 2、生成action_time
                    # action_time_value = fake.date_time_between(start_date=datetime.datetime(2023, 9, 4, 8, 30, 00), end_date=datetime.datetime(2023, 9, 6, 23, 30, 00))
                    action_time_value = fake.date_between(start_date=datetime.date(2023, 9, 4),
                                                          end_date=datetime.date(2023, 9, 6))
                    hour_temp = fake.random_int(min=10, max=22, step=1)
                    if hour_temp < 10:
                        hour_temp2 = '0' + str(hour_temp)
                    else:
                        hour_temp2 = str(hour_temp)
                    min_temp = fake.random_int(min=0, max=59, step=1)
                    if min_temp < 10:
                        min_temp2 = '0' + str(min_temp)
                    else:
                        min_temp2 = str(min_temp)
                    sec_temp = fake.random_int(min=0, max=59, step=1)
                    if sec_temp < 10:
                        sec_temp2 = '0' + str(sec_temp)
                    else:
                        sec_temp2 = str(sec_temp)
                    action_time_value = str(action_time_value) + ' ' + hour_temp2 + ':' + min_temp2 + ':' + sec_temp2
                    # 3、赋值user_id
                    # 4、生成action_type
                    action_type_dict = OrderedDict([('无卡用户+收到资料打开链接', 0.25), ('无', 0.75)])
                    action_type_value = fake.random_element(action_type_dict)
                    # 5、生成amount
                    # 6、生成shop_no
                    if action_type_value != '无':
                        ii = ii + 1
                        action_id.append(ii)
                        action_time.append(action_time_value)
                        user_id.append(user_id_value)
                        action_type.append(action_type_value)
                        amount.append(0)
                        shop_no.append('无')
                    if action_type_value != '无':
                        """  第5步：生成 申请信用卡 的事件数据 """
                        # 1、生成action_id
                        # 2、生成action_time
                        # action_time_value = fake.date_time_between(start_date=datetime.datetime(2023, 9, 7, 8, 30, 00), end_date=datetime.datetime(2023, 9, 10, 23, 30, 00))
                        action_time_value = fake.date_between(start_date=datetime.date(2023, 9, 7),
                                                              end_date=datetime.date(2023, 9, 10))
                        hour_temp = fake.random_int(min=10, max=22, step=1)
                        if hour_temp < 10:
                            hour_temp2 = '0' + str(hour_temp)
                        else:
                            hour_temp2 = str(hour_temp)
                        min_temp = fake.random_int(min=0, max=59, step=1)
                        if min_temp < 10:
                            min_temp2 = '0' + str(min_temp)
                        else:
                            min_temp2 = str(min_temp)
                        sec_temp = fake.random_int(min=0, max=59, step=1)
                        if sec_temp < 10:
                            sec_temp2 = '0' + str(sec_temp)
                        else:
                            sec_temp2 = str(sec_temp)
                        action_time_value = str(
                            action_time_value) + ' ' + hour_temp2 + ':' + min_temp2 + ':' + sec_temp2
                        # 3、赋值user_id
                        # 4、生成action_type
                        action_type_dict = OrderedDict([('无卡用户+申请信用卡', 0.50), ('无', 0.50)])
                        action_type_value = fake.random_element(action_type_dict)
                        # 5、生成amount
                        # 6、生成shop_no
                        if action_type_value != '无':
                            ii = ii + 1
                            action_id.append(ii)
                            action_time.append(action_time_value)
                            user_id.append(user_id_value)
                            action_type.append(action_type_value)
                            amount.append(0)
                            shop_no.append('无')
        """  如果是有卡客户，则根据营销活动类型，需要准备对应的节点事件数据  """
        if have_creditcard == '是':
            temp_dict = OrderedDict([('积分翻倍', 0.30), ('消费满减', 0.30), ('免息分期', 0.40)])
            temp_dict_value = fake.random_element(temp_dict)
            """  如果是 积分翻倍 和 消费满减，则销售流程节点为三个，需要准备对应的节点事件数据  """
            if temp_dict_value in ['积分翻倍','消费满减']:
                """  第1步：生成 发送短信 的事件数据（所有符合条件的用户都发送短信） """
                # 1、生成action_id
                action_id.append(ii)
                # 2、生成action_time
                action_time_value = fake.date_time_between(start_date=datetime.datetime(2023, 9, 1, 9, 30, 00),
                                                           end_date=datetime.datetime(2023, 9, 1, 10, 00, 00))
                action_time.append(action_time_value)
                # 3、赋值user_id
                user_id.append(user_id_value)
                # 4、生成action_type
                action_type.append('有卡用户+'+temp_dict_value + '+发送活动资料')
                # 5、生成amount
                amount.append(0)
                # 6、生成shop_no
                shop_no.append('无')

                """  第2步：生成 收到短信打开链接 的事件数据 """
                # 1、生成action_id
                # 2、生成action_time
                # action_time_value = fake.date_time_between(start_date=datetime.datetime(2023, 9, 2, 9, 30, 00), end_date=datetime.datetime(2023, 9, 4, 23, 30, 00))
                action_time_value = fake.date_between(start_date=datetime.date(2023, 9, 2),
                                                      end_date=datetime.date(2023, 9, 4))
                hour_temp = fake.random_int(min=10, max=22, step=1)
                if hour_temp < 10:
                    hour_temp2 = '0' + str(hour_temp)
                else:
                    hour_temp2 = str(hour_temp)
                min_temp = fake.random_int(min=0, max=59, step=1)
                if min_temp < 10:
                    min_temp2 = '0' + str(min_temp)
                else:
                    min_temp2 = str(min_temp)
                sec_temp = fake.random_int(min=0, max=59, step=1)
                if sec_temp < 10:
                    sec_temp2 = '0' + str(sec_temp)
                else:
                    sec_temp2 = str(sec_temp)
                action_time_value = str(action_time_value) + ' ' + hour_temp2 + ':' + min_temp2 + ':' + sec_temp2
                # 3、赋值user_id
                # 4、生成action_type
                action_type_dict = OrderedDict([('有卡用户+'+temp_dict_value+'+打开活动资料', 0.50), ('无', 0.50)])
                action_type_value = fake.random_element(action_type_dict)
                # 5、生成amount
                # 6、生成shop_no
                if action_type_value != '无':
                    ii = ii + 1
                    action_id.append(ii)
                    action_time.append(action_time_value)
                    user_id.append(user_id_value)
                    action_type.append(action_type_value)
                    amount.append(0)
                    shop_no.append('无')
            if temp_dict_value == '免息分期':
                """  第1步：生成 发送短信 的事件数据（所有符合条件的用户都发送短信） """
                # 1、生成action_id
                action_id.append(ii)
                # 2、生成action_time
                action_time_value = fake.date_time_between(start_date=datetime.datetime(2023, 9, 1, 9, 30, 00),
                                                           end_date=datetime.datetime(2023, 9, 1, 10, 00, 00))
                action_time.append(action_time_value)
                # 3、赋值user_id
                user_id.append(user_id_value)
                # 4、生成action_type
                action_type.append('有卡用户+'+temp_dict_value + '+发送活动资料')
                # 5、生成amount
                amount.append(0)
                # 6、生成shop_no
                shop_no.append('无')

                """  第2步：生成 收到短信打开链接 的事件数据 """
                # 1、生成action_id
                # 2、生成action_time
                # action_time_value = fake.date_time_between(start_date=datetime.datetime(2023, 9, 2, 9, 30, 00), end_date=datetime.datetime(2023, 9, 4, 23, 30, 00))
                action_time_value = fake.date_between(start_date=datetime.date(2023, 9, 2),
                                                      end_date=datetime.date(2023, 9, 4))
                hour_temp = fake.random_int(min=10, max=22, step=1)
                if hour_temp < 10:
                    hour_temp2 = '0' + str(hour_temp)
                else:
                    hour_temp2 = str(hour_temp)
                min_temp = fake.random_int(min=0, max=59, step=1)
                if min_temp < 10:
                    min_temp2 = '0' + str(min_temp)
                else:
                    min_temp2 = str(min_temp)
                sec_temp = fake.random_int(min=0, max=59, step=1)
                if sec_temp < 10:
                    sec_temp2 = '0' + str(sec_temp)
                else:
                    sec_temp2 = str(sec_temp)
                action_time_value = str(action_time_value) + ' ' + hour_temp2 + ':' + min_temp2 + ':' + sec_temp2
                # 3、赋值user_id
                # 4、生成action_type
                action_type_dict = OrderedDict([('有卡用户+'+temp_dict_value+'+打开活动资料', 0.50), ('无', 0.50)])
                action_type_value = fake.random_element(action_type_dict)
                # 5、生成amount
                # 6、生成shop_no
                if action_type_value != '无':
                    ii = ii + 1
                    action_id.append(ii)
                    action_time.append(action_time_value)
                    user_id.append(user_id_value)
                    action_type.append(action_type_value)
                    amount.append(0)
                    shop_no.append('无')
                if action_type_value != '无':
                    """  第3步：生成 申请分期 的事件数据 """
                    # 1、生成action_id
                    # 2、生成action_time
                    # action_time_value = fake.date_time_between(start_date=datetime.datetime(2023, 9, 5, 8, 30, 00), end_date=datetime.datetime(2023, 9, 7, 23, 30, 00))
                    action_time_value = fake.date_between(start_date=datetime.date(2023, 9, 5),
                                                          end_date=datetime.date(2023, 9, 7))
                    hour_temp = fake.random_int(min=10, max=22, step=1)
                    if hour_temp < 10:
                        hour_temp2 = '0' + str(hour_temp)
                    else:
                        hour_temp2 = str(hour_temp)
                    min_temp = fake.random_int(min=0, max=59, step=1)
                    if min_temp < 10:
                        min_temp2 = '0' + str(min_temp)
                    else:
                        min_temp2 = str(min_temp)
                    sec_temp = fake.random_int(min=0, max=59, step=1)
                    if sec_temp < 10:
                        sec_temp2 = '0' + str(sec_temp)
                    else:
                        sec_temp2 = str(sec_temp)
                    action_time_value = str(action_time_value) + ' ' + hour_temp2 + ':' + min_temp2 + ':' + sec_temp2
                    # 3、赋值user_id
                    # 4、生成action_type
                    action_type_dict = OrderedDict([('有卡用户+'+temp_dict_value+'+申请分期', 0.34), ('无', 0.66)])
                    action_type_value = fake.random_element(action_type_dict)
                    # 5、生成amount
                    # 6、生成shop_no
                    if action_type_value != '无':
                        ii = ii + 1
                        action_id.append(ii)
                        action_time.append(action_time_value)
                        user_id.append(user_id_value)
                        action_type.append(action_type_value)
                        amount.append(0)
                        shop_no.append('无')
        data = {
            '事件ID': action_id,
            '事件发生时间': action_time,
            '客户ID': user_id,
            '事件类型': action_type,
            '消费金额': amount,
            '收款方ID': shop_no
        }
    pd.set_option('display.width', None)  # 设置字符显示无限制
    pd.set_option('display.max_rows', None)  # 设置行数显示无限制
    user_actions = pd.DataFrame.from_dict(data)
    # print(user_actions)
    user_actions.to_csv("cgb_user_actions_sales.xlsx", index=True)
    print("\n" + "----------执行结束，万幸----------")

if __name__ == '__main__':
    main()