"""
        自动生成用户事件数据
        Created By：段小飞
        Created Time：2023-09-12
        update Time：2024-04-11
"""
import datetime
import random
import pandas as pd
from datetime import date
from faker import Faker
from tqdm import tqdm
from openpyxl import load_workbook
from collections import OrderedDict

def main():
    fake = Faker('zh-CN')  # 指定用中文
    filename = 'cgb_user.xlsx'  # 定义文件名
    sheetName = 'cgb_user'  # 定义工作表名称

    """  初始化事件属性字段以及字段对应的数据字典  """
    action_id = []      # 1、事件ID
    action_time = []    # 2、事件发生时间
    user_id = []        # 2、客户ID
    action_type = []    # 3、事件类型：逾期还款
    amount = []         # 4、应还款金额
    credit_period = []        # 5、账期

    print("----------1、读取用户信息表----------")
    wb = load_workbook(filename)
    sheet = wb.get_sheet_by_name(sheetName)  # 通过工作表名获取一个工作表对象
    rows = [val for val in sheet.rows]  # 先用列表推导式生产包含每一列中所有单元格的元组的列表
    maxRowNum = sheet.max_row  # 获取工作表中的最大行号
    print("----------2、开始生成数据，祈祷不报错----------")
    for i in tqdm(range(0, maxRowNum)):
        """  1、生成user_id，读取之前生成的用户数据，随机取数，形成用户ID  """
        if i > 0:
            row = rows[i]
            user_id_value = row[0].value  # 获取用户ID
            total_amount_value = float(row[12].value)  # 获取信用卡总额度
            # 获取贷款金额
            loan_amount_value = float(row[15].value)
            if loan_amount_value > 0:
                """  1、生成action_id  """
                action_id_value = fake.random_int(min=10000000, max=99999999, step=1)
                action_id.append(action_id_value)
                """  2、生成action_time  """
                action_time_value = fake.date_between(start_date=date(2024, 1, 1), end_date=date(2024, 5, 31))
                hour_temp = '00'
                min_temp = '00'
                sec_temp = '00'
                action_time_value = str(action_time_value) + ' ' + hour_temp + ':' + min_temp + ':' + sec_temp
                action_time.append(action_time_value)
                """  3、赋值user_id"""
                user_id.append(user_id_value)
                """  4、生成action_type  """
                action_type_dict = OrderedDict([('逾期还款', 0.13), ('正常还款', 0.88)])
                action_type_value = fake.random_element(action_type_dict)
                action_type.append(action_type_value)
                """  5、生成amount  """
                amount_value = round(float(loan_amount_value / 12), 2)
                amount.append(amount_value)
                """  6、生成shop_no  """
                credit_period_dict = OrderedDict([('2024-01', 0.2), ('2024-02', 0.2), ('2024-03', 0.2), ('2024-04', 0.2), ('2024-05', 0.2)])
                credit_period_value = fake.random_element(credit_period_dict)
                credit_period.append(credit_period_value)


        data = {
            '事件ID': action_id,
            '事件发生时间': action_time,
            '客户ID': user_id,
            '用户ID': user_id,
            '事件类型': action_type,
            '应还款金额': amount,
            '还款账期': credit_period
        }
    pd.set_option('display.width', None)  # 设置字符显示无限制
    pd.set_option('display.max_rows', None)  # 设置行数显示无限制
    user_actions = pd.DataFrame.from_dict(data)
    # print(user_actions)
    user_actions.to_csv("cgb_user_repay.xlsx", index=True)
    print("\n" + "----------执行结束，万幸----------")

if __name__ == '__main__':
    main()