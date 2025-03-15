import datetime
from datetime import datetime, timedelta
import random
import pandas as pd
from datetime import date
from faker import Faker
from tqdm import tqdm
from openpyxl import load_workbook
from collections import OrderedDict

def main():
    fake = Faker('zh-CN')  # 指定用中文
    filename = 'cgb_user.xlsx'  # 定义文件名
    sheetName = 'cgb_user'  # 定义工作表名称

    filename2 = 'cgb_user_actions.xlsx'  # 定义文件名
    sheetName2 = 'cgb_user_actions'  # 定义工作表名称

    user_id = []                                    # 1、客户ID
    ssn = []                                        # 2、身份证号（不导出为数据）
    sex = []                                        # 3、性别
    mobile = []                                     # 4、手机号码
    user_name = []                                  # 5、姓名
    credit_no = []                                  # 6、信用卡号
    age = []                                        # 7、用户年龄
    use_amount = []                                 # 8、已用额度
    balance = []                                    # 9、余额
    birthday = []                                   # 10、生日
    channel = []                                    # 11、渠道
    have_creditcard = []                            # 12、是否持有信用卡
    is_installment = []                             # 13、是否分期
    total_amount = []                               # 14、总额度
    yueshouru = []                                  # 15、月收入 存在不同年龄到，不同工作，不同省份地区的月收入不一样的情况要考虑
    creditcard_level = []                           # 16、信用卡级别（普卡、金卡、白金卡、钻石卡、无限卡）
    loan_amount = []                                # 17、贷款金额
    have_house = []                                 # 18、是否有房
    have_car = []                                   # 19、是否有车
    children = []                                   # 20、子女数量
    spreader_id = []                                # 21、传播者ID
    last_shopping_days = []                         # 22、距离最近一次刷卡天数
    num_shopping = []                               # 23、最近30天刷卡次数
    avg_shopping_amount = []                        # 24、最近30天平均刷卡金额

    print("----------1、读取用户信息表----------")
    wb = load_workbook(filename)
    sheet = wb.get_sheet_by_name(sheetName)  # 通过工作表名获取一个工作表对象
    rows = [val for val in sheet.rows]  # 先用列表推导式生产包含每一列中所有单元格的元组的列表
    maxRowNum = sheet.max_row  # 获取工作表中的最大行号
    print("----------2、开始生成数据，祈祷不报错----------")
    for i in tqdm(range(0, maxRowNum)):
        if i > 0:
            row = rows[i]
            # 获取用户ID
            user_id_value = row[0].value
            user_id.append(user_id_value)
            # 3、性别
            sex_value = row[1].value
            sex.append(sex_value)
            # 4、手机号码
            mobile_value = row[2].value
            mobile.append(mobile_value)
            # 5、姓名
            user_name_value = row[3].value
            user_name.append(user_name_value)
            # 6、信用卡号
            credit_no_value = row[4].value
            credit_no.append(credit_no_value)
            # 7、用户年龄
            age_value = row[5].value
            age.append(age_value)
            # 8、已用额度
            use_amount_value = row[6].value
            use_amount.append(use_amount_value)
            # 9、余额
            balance_value = row[7].value
            balance.append(balance_value)
            # 10、生日
            birthday_value = row[8].value
            birthday.append(birthday_value)
            # 11、渠道
            channel_value = row[9].value
            channel.append(channel_value)
            # 12、是否持有信用卡
            have_creditcard_value = row[10].value
            have_creditcard.append(have_creditcard_value)
            # 13、是否分期
            is_installment_value = row[11].value
            is_installment.append(is_installment_value)
            # 14、总额度
            total_amount_value = row[12].value
            total_amount.append(total_amount_value)
            # 15、月收入
            yueshouru_value = row[13].value
            yueshouru.append(yueshouru_value)
            # 16、信用卡级别
            creditcard_level_value = row[14].value
            creditcard_level.append(creditcard_level_value)
            # 17、贷款金额
            loan_amount_value = row[15].value
            loan_amount.append(loan_amount_value)
            # 18、是否有房
            have_house_value = row[16].value
            have_house.append(have_house_value)
            # 19、是否有车
            have_car_value = row[17].value
            have_car.append(have_car_value)
            # 20、子女数量
            children_value = row[18].value
            children.append(children_value)
            # 21、传播者ID
            spreader_id_value = row[19].value
            spreader_id.append(spreader_id_value)

            print("----------3、读取用户事件行为信息表----------")
            wb2 = load_workbook(filename2)
            sheet2 = wb2.get_sheet_by_name(sheetName2)  # 通过工作表名获取一个工作表对象
            rows2 = [val for val in sheet2.rows]  # 先用列表推导式生产包含每一列中所有单元格的元组的列表
            maxRowNum2 = sheet2.max_row  # 获取工作表中的最大行号
            print("----------4、开始逐个用户统计消费RFM相关关键数据----------")
            num_shopping_temp = 0
            shopping_amount_temp = 0
            shopping_days_temp = datetime.strptime('2024-01-01', "%Y-%m-%d")
            # 每一个用户，都到事件表里面逐行去查询该用户的事件信息，然后进行统计
            for j in tqdm(range(1, maxRowNum2)):
                row2 = rows2[j]
                # 获得事件信息的user_id
                user_id_action_value = row2[2].value
                # print('user_id:',user_id_action_value)
                # 判断该事件的user_id是不是和当前用户的user_id一致，也就是是不是一个用户
                if user_id_action_value == user_id_value:
                    # print('cation_id:',row2[0].value)
                    action_time = row2[1].value
                    if action_time > shopping_days_temp:
                        shopping_days_temp = action_time
                    if within_30_days(action_time):
                        # 消费次数增加
                        num_shopping_temp = num_shopping_temp + 1
                        # 累计消费金额
                        shopping_amount_temp = shopping_amount_temp + row2[4].value

            # 22、距离最近一次刷卡天数
            current_date = datetime.now()
            last_shopping_days_value = (current_date - shopping_days_temp).days
            last_shopping_days.append(last_shopping_days_value)
            # 23、最近30天刷卡次数
            num_shopping_value = num_shopping_temp
            num_shopping.append(num_shopping_value)
            # 24、最近30天平均刷卡金额
            if num_shopping_temp > 0:
                avg_shopping_amount_value = round(shopping_amount_temp/num_shopping_temp, 2)
            else:
                avg_shopping_amount_value = 0
            avg_shopping_amount.append(avg_shopping_amount_value)
            print('用户ID: ', user_id_value, '最近30天刷卡次数：', num_shopping_temp, '最近30天平均刷卡金额：', avg_shopping_amount_value, '距离最近一次刷卡天数：', last_shopping_days_value)



        data = {
            '用户ID': user_id,
            '性别': sex,
            '手机号码': mobile,
            '姓名': user_name,
            '信用卡号': credit_no,
            '年龄': age,
            '已用额度': use_amount,
            '余额': balance,
            '生日': birthday,
            '渠道': channel,
            '是否持有信用卡': have_creditcard,
            '是否分期': is_installment,
            '总额度': total_amount,
            '月收入': yueshouru,
            '信用卡级别': creditcard_level,
            '贷款金额': loan_amount,
            '是否有房': have_house,
            '是否有车': have_car,
            '子女数量': children,
            '传播者ID': spreader_id,
            '距离最近一次刷卡天数': last_shopping_days,
            '最近30天刷卡次数': num_shopping,
            '最近30天平均刷卡金额': avg_shopping_amount
        }
    pd.set_option('display.width', None)  # 设置字符显示无限制
    pd.set_option('display.max_rows', None)  # 设置行数显示无限制
    user_actions = pd.DataFrame.from_dict(data)
    # print(user_actions)
    user_actions.to_csv("用户信息-计算后.xlsx", index=True)
    print("\n" + "----------执行结束，万幸----------")

def within_30_days(target_time):
    now = datetime.now()
    thirty_days_ago = now - timedelta(days=30)
    return target_time > thirty_days_ago and target_time <= now

def days_between_dates(date1, date2):
    date1 = datetime.strptime(date1, "%Y-%m-%d")
    date2 = datetime.strptime(date2, "%Y-%m-%d")
    return (date2 - date1).days


if __name__ == '__main__':
    main()