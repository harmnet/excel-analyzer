"""
        自动生成用户事件数据
        Created By：段小飞
        Created Time：2023-09-12
        update Time：2024-04-11
"""
import datetime
import random
import pandas as pd
from datetime import date
from faker import Faker
from tqdm import tqdm
from openpyxl import load_workbook
from collections import OrderedDict

def main():
    fake = Faker('zh-CN')  # 指定用中文
    filename = '用户信息-计算后.xlsx'  # 定义文件名
    sheetName = '用户信息-计算后'  # 定义工作表名称

    """  初始化事件属性字段以及字段对应的数据字典  """
    action_id = []      # 1、事件ID
    action_time = []    # 2、事件发生时间
    user_id = []        # 3、客户ID
    action_type = []    # 4、事件类型：传播裂变

    print("----------1、读取用户信息表----------")
    wb = load_workbook(filename)
    sheet = wb.get_sheet_by_name(sheetName)  # 通过工作表名获取一个工作表对象
    rows = [val for val in sheet.rows]  # 先用列表推导式生产包含每一列中所有单元格的元组的列表
    maxRowNum = sheet.max_row  # 获取工作表中的最大行号
    print("----------2、开始生成数据，祈祷不报错----------")
    for i in tqdm(range(0, maxRowNum)):
        """  1、生成user_id，读取之前生成的用户数据，随机取数，形成用户ID  """
        if i > 0:
            row = rows[i]
            user_id_value = row[0].value  # 获取用户ID
            # 传播者ID
            spreader_id_value = float(row[19].value)
            if spreader_id_value > 0:
                """  1、生成action_id  """
                action_id_value = fake.random_int(min=10000000, max=99999999, step=1)
                action_id.append(action_id_value)
                """  2、生成action_time  """
                action_time_value = fake.date_between(start_date=date(2024, 1, 1), end_date=date(2024, 5, 31))
                hour_temp = fake.random_int(min=9, max=22, step=1)
                if hour_temp < 10:
                    hour_temp2 = '0' + str(hour_temp)
                else:
                    hour_temp2 = str(hour_temp)
                min_temp = fake.random_int(min=0, max=59, step=1)
                if min_temp < 10:
                    min_temp2 = '0' + str(min_temp)
                else:
                    min_temp2 = str(min_temp)
                sec_temp = fake.random_int(min=0, max=59, step=1)
                if sec_temp < 10:
                    sec_temp2 = '0' + str(sec_temp)
                else:
                    sec_temp2 = str(sec_temp)
                action_time_value = str(action_time_value) + ' ' + hour_temp2 + ':' + min_temp2 + ':' + sec_temp2
                action_time.append(action_time_value)
                """  3、赋值user_id"""
                user_id.append(user_id_value)
                """  4、生成action_type  """
                action_type_value = '传播裂变'
                action_type.append(action_type_value)



        data = {
            '事件ID': action_id,
            '事件发生时间': action_time,
            '客户ID': user_id,
            '用户ID': user_id,
            '事件类型': action_type
        }
    pd.set_option('display.width', None)  # 设置字符显示无限制
    pd.set_option('display.max_rows', None)  # 设置行数显示无限制
    user_actions = pd.DataFrame.from_dict(data)
    # print(user_actions)
    user_actions.to_csv("cgb_user_split.xlsx", index=True)
    print("\n" + "----------执行结束，万幸----------")

if __name__ == '__main__':
    main()