"""
        自动生成用户事件数据
        Created By：段小飞
        Created Time：2023-09-12
        update Time：2024-04-11
"""
import datetime
import random
import pandas as pd
from datetime import date
from faker import Faker
from tqdm import tqdm
from openpyxl import load_workbook
from collections import OrderedDict

def main():
    fake = Faker('zh-CN')  # 指定用中文
    num = 3000  # 初始化模拟数据生成数量
    filename = 'cgb_user.xlsx'  # 定义文件名
    sheetName = 'cgb_user'  # 定义工作表名称

    """  初始化事件属性字段以及字段对应的数据字典  """
    action_id = []      # 1、事件ID
    action_time = []    # 2、事件发生时间
    user_id = []        # 2、客户ID
    action_type = []    # 3、事件类型：餐饮、大额消费、电商、购物、旅游、其他
    amount = []         # 4、消费金额
    shop_no = []        # 5、收款方ID

    print("----------1、读取用户信息表----------")
    wb = load_workbook(filename)
    sheet = wb.get_sheet_by_name(sheetName)  # 通过工作表名获取一个工作表对象
    rows = [val for val in sheet.rows]  # 先用列表推导式生产包含每一列中所有单元格的元组的列表
    maxRowNum = sheet.max_row  # 获取工作表中的最大行号
    print("----------2、开始生成数据，祈祷不报错----------")
    ii = 0
    for i in tqdm(range(0, num)):
        """  1、生成user_id，读取之前生成的用户数据，随机取数，形成用户ID  """
        indexs = random.sample(range(1, maxRowNum - 1), 1)  # 获取0到最大行号中的1个整数
        for index in indexs:
            row = rows[index]
            user_id_value = row[0].value  # 获取用户ID
            total_amount_value = float(row[12].value)  # 获取信用卡总额度
        random_num = fake.random_int(min=1, max=20)

        for j in tqdm(range(0, random_num)):
            ii = ii +1
            """  1、生成action_id  """
            action_id_value = fake.random_int(min=10000000, max=99999999, step=1)
            action_id.append(ii)
            """  2、生成action_time  """
            action_time_value = fake.date_between(start_date=date(2024, 1, 1), end_date=date(2024, 5, 31))
            hour_temp = fake.random_int(min=9, max=22, step=1)
            if hour_temp < 10:
                hour_temp2 = '0' + str(hour_temp)
            else: hour_temp2 = str(hour_temp)
            min_temp = fake.random_int(min=0, max=59, step=1)
            if min_temp < 10:
                min_temp2 = '0' + str(min_temp)
            else:
                min_temp2 = str(min_temp)
            sec_temp = fake.random_int(min=0, max=59, step=1)
            if sec_temp < 10:
                sec_temp2 = '0' + str(sec_temp)
            else:
                sec_temp2 = str(sec_temp)
            action_time_value = str(action_time_value) + ' ' + hour_temp2 + ':' + min_temp2 + ':' + sec_temp2

            action_time.append(action_time_value)
            """  3、赋值user_id"""
            user_id.append(user_id_value)
            """  4、生成action_type  """
            action_type_dict = OrderedDict([('餐饮', 0.23), ('大额消费', 0.05), ('电商', 0.26), ('购物', 0.13), ('旅游', 0.11), ('出行', 0.15), ('其他', 0.07)])
            action_type_value = fake.random_element(action_type_dict)
            action_type.append(action_type_value)
            """  5、生成amount  """
            if 100000 < total_amount_value <= 150000:
                if action_type_value == '餐饮':
                    amount_value = fake.random_int(min=30000, max=100000, step=1)/100
                elif action_type_value == '大额消费':
                    amount_value = fake.random_int(min=1000000, max=10000000, step=1) / 100
                elif action_type_value == '电商':
                    amount_value = fake.random_int(min=30000, max=300000, step=1) / 100
                elif action_type_value == '购物':
                    amount_value = fake.random_int(min=30000, max=500000, step=1) / 100
                elif action_type_value == '旅游':
                    amount_value = fake.random_int(min=500000, max=2000000, step=1) / 100
                elif action_type_value == '其他':
                    amount_value = fake.random_int(min=30000, max=100000, step=1) / 100
                elif action_type_value == '出行':
                    amount_value = fake.random_int(min=5000, max=300000, step=1) / 100
            elif 50000 < total_amount_value <= 100000:
                if action_type_value == '餐饮':
                    amount_value = fake.random_int(min=20000, max=80000, step=1) / 100
                elif action_type_value == '大额消费':
                    amount_value = fake.random_int(min=500000, max=5000000, step=1) / 100
                elif action_type_value == '电商':
                    amount_value = fake.random_int(min=20000, max=200000, step=1) / 100
                elif action_type_value == '购物':
                    amount_value = fake.random_int(min=10000, max=300000, step=1) / 100
                elif action_type_value == '旅游':
                    amount_value = fake.random_int(min=300000, max=1000000, step=1) / 100
                elif action_type_value == '其他':
                    amount_value = fake.random_int(min=10000, max=100000, step=1) / 100
                elif action_type_value == '出行':
                    amount_value = fake.random_int(min=5000, max=300000, step=1) / 100
            elif 20000 < total_amount_value <= 50000:
                if action_type_value == '餐饮':
                    amount_value = fake.random_int(min=10000, max=50000, step=1) / 100
                elif action_type_value == '大额消费':
                    amount_value = fake.random_int(min=300000, max=1000000, step=1) / 100
                elif action_type_value == '电商':
                    amount_value = fake.random_int(min=10000, max=100000, step=1) / 100
                elif action_type_value == '购物':
                    amount_value = fake.random_int(min=10000, max=300000, step=1) / 100
                elif action_type_value == '旅游':
                    amount_value = fake.random_int(min=300000, max=1000000, step=1) / 100
                elif action_type_value == '其他':
                    amount_value = fake.random_int(min=10000, max=500000, step=1) / 100
                elif action_type_value == '出行':
                    amount_value = fake.random_int(min=5000, max=300000, step=1) / 100
            elif 10000 < total_amount_value <= 20000:
                if action_type_value == '餐饮':
                    amount_value = fake.random_int(min=50000, max=50000, step=1) / 100
                elif action_type_value == '大额消费':
                    amount_value = fake.random_int(min=200000, max=500000, step=1) / 100
                elif action_type_value == '电商':
                    amount_value = fake.random_int(min=50000, max=50000, step=1) / 100
                elif action_type_value == '购物':
                    amount_value = fake.random_int(min=5000, max=50000, step=1) / 100
                elif action_type_value == '旅游':
                    amount_value = fake.random_int(min=100000, max=300000, step=1) / 100
                elif action_type_value == '其他':
                    amount_value = fake.random_int(min=50000, max=50000, step=1) / 100
                elif action_type_value == '出行':
                    amount_value = fake.random_int(min=5000, max=300000, step=1) / 100
            elif 5000 < total_amount_value <= 10000:
                if action_type_value == '餐饮':
                    amount_value = fake.random_int(min=10000, max=30000, step=1) / 100
                elif action_type_value == '大额消费':
                    amount_value = fake.random_int(min=100000, max=300000, step=1) / 100
                elif action_type_value == '电商':
                    amount_value = fake.random_int(min=2000, max=30000, step=1) / 100
                elif action_type_value == '购物':
                    amount_value = fake.random_int(min=2000, max=30000, step=1) / 100
                elif action_type_value == '旅游':
                    amount_value = fake.random_int(min=50000, max=100000, step=1) / 100
                elif action_type_value == '其他':
                    amount_value = fake.random_int(min=2000, max=30000, step=1) / 100
                elif action_type_value == '出行':
                    amount_value = fake.random_int(min=5000, max=300000, step=1) / 100
            elif 2000 < total_amount_value <= 5000:
                if action_type_value == '餐饮':
                    amount_value = fake.random_int(min=5000, max=20000, step=1) / 100
                elif action_type_value == '大额消费':
                    amount_value = fake.random_int(min=50000, max=100000, step=1) / 100
                elif action_type_value == '电商':
                    amount_value = fake.random_int(min=1000, max=20000, step=1) / 100
                elif action_type_value == '购物':
                    amount_value = fake.random_int(min=1000, max=20000, step=1) / 100
                elif action_type_value == '旅游':
                    amount_value = fake.random_int(min=10000, max=80000, step=1) / 100
                elif action_type_value == '其他':
                    amount_value = fake.random_int(min=1000, max=20000, step=1) / 100
                elif action_type_value == '出行':
                    amount_value = fake.random_int(min=5000, max=300000, step=1) / 100
            elif total_amount_value <= 2000:
                if action_type_value == '餐饮':
                    amount_value = fake.random_int(min=1000, max=10000, step=1) / 100
                elif action_type_value == '大额消费':
                    amount_value = fake.random_int(min=30000, max=50000, step=1) / 100
                elif action_type_value == '电商':
                    amount_value = fake.random_int(min=500, max=10000, step=1) / 100
                elif action_type_value == '购物':
                    amount_value = fake.random_int(min=500, max=10000, step=1) / 100
                elif action_type_value == '旅游':
                    amount_value = fake.random_int(min=10000, max=50000, step=1) / 100
                elif action_type_value == '其他':
                    amount_value = fake.random_int(min=500, max=10000, step=1) / 100
                elif action_type_value == '出行':
                    amount_value = fake.random_int(min=5000, max=100000, step=1) / 100
            amount.append(amount_value)
            """  6、生成shop_no  """
            shop_no_value = fake.random_int(min=100000, max=999999, step=1)
            shop_no.append(shop_no_value)


        data = {
            '事件ID': action_id,
            '事件发生时间': action_time,
            '客户ID': user_id,
            '事件类型': action_type,
            '消费金额': amount,
            '收款方ID': shop_no
        }
    pd.set_option('display.width', None)  # 设置字符显示无限制
    pd.set_option('display.max_rows', None)  # 设置行数显示无限制
    user_actions = pd.DataFrame.from_dict(data)
    # print(user_actions)
    user_actions.to_csv("cgb_user_actions.xlsx", index=True)
    print("\n" + "----------执行结束，万幸----------")

if __name__ == '__main__':
    main()