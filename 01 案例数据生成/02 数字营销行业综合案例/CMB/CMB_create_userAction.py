"""
        自动生成用户事件数据
        案例：CMB
        Created By：xujian
        Created Time：2022-11-28
"""
import random
import pandas as pd
from datetime import date
from faker import Faker
from tqdm import tqdm
from openpyxl import load_workbook
from collections import OrderedDict

def main():
    fake = Faker('zh-CN')  # 指定用中文
    num = 10000  # 初始化模拟数据生成数量
    filename = 'CMB_user.xlsx'  # 定义文件名
    sheetName = 'CMB_user'  # 定义工作表名称
    """  初始化事件属性字段以及字段对应的数据字典  """
    action_id = []      # 1、用户事件ID
    user_id = []        # 2、客户ID
    action_type = []    # 3、事件类型
    action_detail = []  # 4、事件动作
    action_time = []    # 5、事件时间
    amount = []         # 6、购买金额
    print("----------1、读取用户信息表----------")
    wb = load_workbook(filename)
    sheet = wb.get_sheet_by_name(sheetName)  # 通过工作表名获取一个工作表对象
    rows = [val for val in sheet.rows]  # 先用列表推导式生产包含每一列中所有单元格的元组的列表
    maxRowNum = sheet.max_row  # 获取工作表中的最大行号
    print("----------2、开始生成数据，祈祷不报错----------")
    for i in tqdm(range(0, num)):
        """  1、生成action_id  """
        action_id_value = fake.random_int(min=10000000, max=99999900, step=1)
        action_id.append(action_id_value)
        """  2、生成user_id，读取之前生成的用户数据，随机取数，形成用户ID  """
        indexs = random.sample(range(0, maxRowNum - 1), 1) # 获取0到最大行号中的1个整数
        for index in indexs:
            row = rows[index]
            user_id_value = row[0].value
        user_id.append(user_id_value)
        """  3、生成action_type  """
        action_type_dict = OrderedDict([('关注公众号', 0.35), ('点击文章', 0.30), ('打开小程序', 0.25), ('下单购买', 0.15)])
        action_type_value = fake.random_element(action_type_dict)
        action_type.append(action_type_value)
        """  4、生成action_detail  """
        if action_type_value == "关注公众号":
            action_detail_dict = OrderedDict([('引商银行公众号', 1)])
            action_detail_value = fake.random_element(action_detail_dict)
            action_detail.append(action_detail_value)
        elif action_type_value == "点击文章":
            action_detail_dict = OrderedDict([('产品介绍文章一', 0.35), ('产品介绍文章二', 0.20), ('产品介绍文章三', 0.25), ('产品介绍文章四', 0.20)])
            action_detail_value = fake.random_element(action_detail_dict)
            action_detail.append(action_detail_value)
        elif action_type_value == "打开小程序":
            action_detail_dict = OrderedDict([('打开小程序', 0.85), ('未打开小程序', 0.15)])
            action_detail_value = fake.random_element(action_detail_dict)
            action_detail.append(action_detail_value)
        elif action_type_value == "下单购买":
            action_detail_dict = OrderedDict([('下单购买', 0.90), ('未下单购买', 0.10)])
            action_detail_value = fake.random_element(action_detail_dict)
            action_detail.append(action_detail_value)
        """  5、生成action_time  """
        action_time_value = fake.date_time_between(start_date=date(2023, 11, 1), end_date=date(2023, 11, 30))
        action_time.append(action_time_value)
        """  5、生成amount  """
        if action_detail_value == '下单购买':
            amount_dict = OrderedDict([('1万以下', 0.00), ('1~5万', 0.00), ('5~10万', 0.00), ('10~20万', 0.00), ('20~50万', 0.10), ('50~100万', 0.55),
             ('100~200万', 0.25), ('200万以上', 0.10)])  # 数据字典：购买金额
            amount_value = fake.random_element(amount_dict)
            amount.append(amount_value)
        else:
            amount_value = 0
            amount.append(amount_value)

            data = {
            '用户事件ID': action_id,
            '客户ID': user_id,
            '事件类型': action_type,
            '事件动作': action_detail,
            '事件时间': action_time,
            '购买金额': amount
        }
    pd.set_option('display.width', None)  # 设置字符显示无限制
    pd.set_option('display.max_rows', None)  # 设置行数显示无限制
    user_actions = pd.DataFrame.from_dict(data)
    # print(user_actions)
    user_actions.to_csv("CMB_user_actions.xlsx", index=True)
    print("\n" + "----------执行结束，万幸----------")

if __name__ == '__main__':
    main()